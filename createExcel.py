from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment


def createExcel(name, bvi, kvota, budjet, max_b, min_p):
    data = [
        [name, "", "", "", "", f"Отчет создан {datetime.now().strftime('%d.%m.%y %H:%M:%S')}"],
        ["Проходной балл на бюджет:", "", (min_p)],
        ["Занятых мест по квоте:", "", (len(kvota))],
        ["Занятых БВИ:", "", (len(bvi))],
        ["Занятых остальнысм:", "", (max_b - len(kvota) - len(bvi))],
        ["№", "СНИЛС", "Особые права", "Баллы ЕГЭ", "", "", "", "Оригинал аттестата", "Приоритет"],
        ["", "", "БВИ/Квота", "Математика", "Русский язык", "Информатика", "Сумма"],
        ["БВИ"]
    ]



    for i in bvi:

        data.append([(i["number"]), i["snils"], (i["bvi"]), (i["math"]),(i["russian"]),(i["informatics"]),(i["summ"]), i["orig"], i["prioritet_inoe"]])
    data.append(["КВОТА"])
    for i in kvota:

        data.append([(i["number"]), i["snils"], (i["bvi"]), (i["math"]),(i["russian"]),(i["informatics"]),(i["summ"]), i["orig"], i["prioritet_inoe"]])
    data.append(["Бюджет"])
    for i in budjet:

        data.append([(i["number"]), i["snils"], (i["bvi"]), (i["math"]),(i["russian"]),(i["informatics"]),(i["summ"]), i["orig"], i["prioritet_inoe"]])


    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.column_dimensions['A'].width = 7
    sheet.column_dimensions['B'].width = 20
    for row_data in data:
        sheet.append(row_data)
    sheet.freeze_panes = sheet["A8"]
    sheet.merge_cells("D6:G6")
    sheet.merge_cells("A6:A7")
    sheet.merge_cells("B6:B7")
    sheet.merge_cells("C6:C7")
    sheet.merge_cells("H6:H7")
    sheet.merge_cells("I6:I7")
    sheet.merge_cells("A8:I8")
    sheet["C6"].alignment = Alignment(wrap_text=True)
    sheet["H6"].alignment = Alignment(wrap_text=True)
    sheet["I6"].alignment = Alignment(wrap_text=True)
    sheet["A8"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["D6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(f"A{8 + 1 + len(bvi)}:I{8 + 1 + len(bvi)}")
    sheet[f"A{8 + 1 + len(bvi)}"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(f"A{8 + 1 + len(bvi) + 1+  len(kvota)}:I{8 + 1 + len(bvi) + 1+  len(kvota)}")
    sheet[f"A{8 + 1 + len(bvi) + 1+  len(kvota)}"].alignment = Alignment(horizontal="center", vertical="center")
    # Create a new sheet


    # Write the array data to the sheet


    # Save the workbook to a file
    workbook.save('xlsx/output__pmi.xlsx')
    print("FILE UPDATED")