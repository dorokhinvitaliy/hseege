import math
import string

import pandas as pd
import openpyxl
def parseExcel(vuz, napr, prior, attestat):
    # Import pandas

    # Give the location of the file
    path = f"loads/{vuz}_{napr}.xlsx"

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    mx = sheet_obj.max_row
    max_budj = sheet_obj.cell(row=8, column=6).value
    verion_date = sheet_obj.cell(row=5, column=6).value
    name = sheet_obj.cell(row=2, column=1).value
    print(max_budj)
    cell_obj = sheet_obj['A16': 'V'+str(mx)]
    array = []

    headers = ["id", "snils", "bvi", "bvi_reason",
               "osobkvota", "celkvota", "otdkvota", "prioritet_cell", "prioritet_inoe", "plat_prioritet",
               "math", "informatics", "russian", "IND", 'summ', 'mesto', 'orig', 'soglasie', 'preimus', 'obsh', 'vozvrat']

    for row in cell_obj:
        vals = [(0 if x.value is None else x.value)  for index,x in enumerate(row) if index not in [2,5,6]]
        obj = dict(zip(headers,vals))
        array.append(obj)
        #print(obj)
    #array = sorted(list(filter(lambda x: x["bvi"]=="Да", array)), key=lambda x: x["summ"], reverse=True) + sorted(list(filter(lambda x: x["bvi"]=="Нет", array)), key=lambda x: x["summ"], reverse=True)
    #result_arr = []
    array = sorted(array, key=lambda x: -int(x["summ"]))
    bvi = []
    osobkvota = []
    celkvota = []
    otdkvota = []
    budj = []

    for i in range(len(array)):
        if prior < (0 if isinstance(array[i]["prioritet_inoe"], str) else array[i]["prioritet_inoe"]): continue
        if attestat != 0:
            if attestat == 1:
                if array[i]["orig"] == "Нет": continue
            else:
                if array[i]["orig"] == "Да": continue
        if array[i]["bvi"] == "Да":
            bvi.append(array[i])
        elif array[i]["osobkvota"] == "Да" and len(osobkvota) < math.ceil(0.1*max_budj):
            array[i]["kvota"] = "Особ."
            osobkvota.append(array[i])
        elif array[i]["celkvota"] == "Да" and len(celkvota) < math.ceil(0.1*max_budj):
            array[i]["kvota"] = "Цел."
            celkvota.append(array[i])
        elif array[i]["otdkvota"] == "Да" and len(otdkvota) < math.ceil(0.1*max_budj):
            array[i]["kvota"] = "Отд."
            otdkvota.append(array[i])
        else:
            budj.append(array[i])
    #new_array = bvi+osobkvota+celkvota+otdkvota+budj
    cnt = 0
    prohod = 0
    for mas in [bvi, osobkvota, celkvota, otdkvota, budj]:
        for i in range(len(mas)):
            cnt+=1
            mas[i]["number"] = cnt
            if cnt==max_budj:
                prohod = mas[i]["summ"]

    """for i in range(len(bvi)):
        cnt+=1
        bvi[i]["number"] = cnt
        
        
    for i in range(len(osobkvota)):
        osobkvota[i]["number"] = i+1
        
        
    for i in range(len(bvi)):
        bvi[i]["number"] = i+1
        
        
    for i in range(len(bvi)):
        bvi[i]["number"] = i+1
        
        
    for i in range(len(bvi)):
        bvi[i]["number"] = i+1
        
    """

    return ([bvi, osobkvota+celkvota+otdkvota, budj, bvi+osobkvota+celkvota+otdkvota+budj],max_budj,prohod,name,verion_date)
